# /// script
# requires-python = ">=3.11"
# dependencies = ["requests", "openpyxl"]
# ///
"""¿Se puede obtener la ficha ampliada de vivienda por el API v1?

El portal nuevo (idg) acepta ``{"mara":"2024","codigos":[...],"vivienda":true}`` y
devuelve una segunda ficha con materiales de construcción, hacinamiento y tipo de
hogar (34 campos). En el API v1 el flag ``vivienda`` parece ignorarse. Este
sondeo prueba variantes del cuerpo para descartar que falte otro parámetro.

Compara cada respuesta contra la ficha base por hash de contenido de celdas, así
un XLSX distinto solo por metadatos del zip no cuenta como diferencia.

Uso:  uv run data-raw/fichas/sondeo_vivienda.py [codigo]
"""

import hashlib
import io
import sys
import warnings

import openpyxl

from ine_api import ClienteINE, ErrorAPI

warnings.filterwarnings("ignore")

CODIGO = sys.argv[1] if len(sys.argv) > 1 else "00417298575-A"

VARIANTES = [
    ("base (referencia)", {"codigos": [CODIGO]}),
    ("vivienda:true", {"codigos": [CODIGO], "vivienda": True}),
    ("vivienda + mara str", {"codigos": [CODIGO], "vivienda": True, "mara": "2024"}),
    ("vivienda + mara int", {"codigos": [CODIGO], "vivienda": True, "mara": 2024}),
    ("vivienda + tipo cpv", {"codigos": [CODIGO], "vivienda": True, "mara": 2024, "tipo": "cpv"}),
    ("tipo=vivienda", {"codigos": [CODIGO], "tipo": "vivienda"}),
    ("ficha=vivienda", {"codigos": [CODIGO], "ficha": "vivienda"}),
    ("esVivienda:true", {"codigos": [CODIGO], "esVivienda": True}),
]


def huella(contenido: bytes) -> tuple[str, int, list[str]]:
    """Hash del contenido real de las celdas, no del archivo."""
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    trozos, n = [], 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None and str(c.value).strip():
                    trozos.append(f"{ws.title}!{c.coordinate}={c.value}")
                    n += 1
    return hashlib.sha256("\n".join(trozos).encode()).hexdigest()[:16], n, wb.sheetnames


def main() -> None:
    cli = ClienteINE()
    ref = None
    for nombre, cuerpo in VARIANTES:
        try:
            b = cli._post("/generar-excel", cuerpo, binario=True)
        except ErrorAPI as e:
            print(f"{nombre:24} -> ERROR {str(e)[:90]}")
            continue
        h, n, hojas = huella(b)
        if ref is None:
            ref = h
            print(f"{nombre:24} -> {len(b):>8,} B  celdas={n}  hash={h}  hojas={hojas}")
        else:
            marca = "IGUAL a la base" if h == ref else ">>> DISTINTA <<<"
            print(f"{nombre:24} -> {len(b):>8,} B  celdas={n}  hash={h}  {marca}")


if __name__ == "__main__":
    main()
